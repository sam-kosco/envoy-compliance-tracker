"""
crosswinds_generate_data.py
============================
Called by crosswinds_refresh.yml with the PAYLOAD env var set to a JSON
array of inspections from Power Automate (last 7 days from SafetyCulture).

What this script does:
  1. Downloads Crosswinds Debriefs.xlsx from SharePoint
  2. Upserts the incoming inspections:
       - If audit_id already exists in the table → update the row
         (covers edits to a SafetyCulture inspection within the PA window)
       - If audit_id is new → append a new row
  3. Re-uploads the updated Excel to SharePoint
  4. Reads the FULL table (all history) to calculate rolling 7-day compliance
  5. Writes crosswinds_data.json for the GitHub Pages dashboard

SharePoint table columns (Crosswind_Debriefs):
  Date | Name | Location | Tail | Audit ID | Template ID | Report Link

Environment variables (GitHub Secrets):
  TENANT_ID, CLIENT_ID, CLIENT_SECRET
"""

import os, json, sys, io, requests
from datetime import datetime, timezone, date, timedelta
from openpyxl import load_workbook

# ── CONSTANTS ────────────────────────────────────────────────────────────────
TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

DRIVE_ID  = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"
FILE_PATH = "Power Flows/Debriefs/Crosswinds Debriefs.xlsx"
TABLE_NAME = "Crosswind_Debriefs"

TAILS = [
    "N351DC","N383CA","N478DC","N2322Y","N536DC","N723AG","N830BS",
    "N390JA","N727MZ","N705Q","N238E","N154DS","N70KK","N633DS",
    "N582DC","N572CW","N82ZZ","N618DC","N595GL","N267DC","N150PT",
    "N557DS","N942RM","N599DC","N11ZM","N625DC","N627DC","N419JS",
    "N37EE","N90FF","N60VU","N518MA","N123TV","N321JE","N84VV",
    "N568DC","N58HH","N141DK","N29FT","N650KN","N714KJ"
]

WINDOW_DAYS = 7   # rolling compliance window
SOON_DAYS   = 4   # cleans this week warning threshold
SHOW_HISTORY = 5  # recent inspections shown in detail panel

# Column order in the Excel table (1-indexed position in the worksheet)
COL = {
    "Date":        1,
    "Name":        2,
    "Location":    3,
    "Tail":        4,
    "Audit ID":    5,
    "Template ID": 6,
    "Report Link": 7,
}


# ── GRAPH API HELPERS ─────────────────────────────────────────────────────────
def get_token():
    print("Acquiring Graph API token...")
    r = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={
            "grant_type":    "client_credentials",
            "client_id":     CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "scope":         "https://graph.microsoft.com/.default",
        }
    )
    r.raise_for_status()
    print("  Token acquired.")
    return r.json()["access_token"]


def download_excel(token):
    encoded = FILE_PATH.replace(" ", "%20")
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{encoded}:/content"
    print(f"Downloading: {FILE_PATH}")
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    print(f"  Downloaded {len(r.content):,} bytes")
    return io.BytesIO(r.content)


def upload_excel(token, buffer):
    encoded = FILE_PATH.replace(" ", "%20")
    url = f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{encoded}:/content"
    print("Uploading updated Excel...")
    r = requests.put(
        url,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
        data=buffer.getvalue()
    )
    r.raise_for_status()
    print(f"  Uploaded successfully (status {r.status_code})")


# ── EXCEL HELPERS ─────────────────────────────────────────────────────────────
def find_table_sheet(wb):
    """Find the sheet containing the Crosswind_Debriefs table."""
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            if tbl.name == TABLE_NAME:
                return ws, tbl
    raise ValueError(f"Table '{TABLE_NAME}' not found in workbook")


def read_table(ws, tbl):
    """Read the table into a list of dicts, keyed by Audit ID."""
    ref = tbl.ref  # e.g. "A1:G100"
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(ref)

    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row,
                              min_col=min_col, max_col=max_col, values_only=True))
    if not rows:
        return {}, min_row, min_col, max_row

    # First row is header
    headers = [str(h).strip() if h else "" for h in rows[0]]
    data = {}
    for row in rows[1:]:
        if not any(row):
            continue
        rec = dict(zip(headers, row))
        aid = str(rec.get("Audit ID") or "").strip()
        if aid:
            data[aid] = rec

    return data, min_row, min_col, max_row


def upsert_records(ws, tbl, existing, incoming):
    """
    Upsert incoming records into the worksheet.
    - existing: dict of {audit_id: row_dict} already in the sheet
    - incoming: list of dicts from Power Automate payload
    Returns count of inserted and updated rows.
    """
    from openpyxl.utils import range_boundaries

    ref = tbl.ref
    min_col, min_row, max_col, max_row = range_boundaries(ref)

    # Build a map of audit_id → row number for existing records
    audit_to_row = {}
    for row_idx in range(min_row + 1, max_row + 1):
        aid_cell = ws.cell(row=row_idx, column=min_col + COL["Audit ID"] - 1)
        if aid_cell.value:
            audit_to_row[str(aid_cell.value).strip()] = row_idx

    inserted = 0
    updated  = 0

    for insp in incoming:
        audit_id = (insp.get("audit_id") or "").strip()
        if not audit_id:
            continue

        # Parse date — store as date object so Excel formats it correctly
        date_val = insp.get("date", "")
        try:
            dt = datetime.fromisoformat(date_val.replace("Z", "+00:00"))
            date_obj = dt.date()
        except Exception:
            date_obj = None

        row_values = [
            date_obj,
            insp.get("tech", ""),
            insp.get("location", ""),
            (insp.get("tail") or "").strip().upper(),
            audit_id,
            insp.get("template_id", ""),
            insp.get("report_url", ""),
        ]

        if audit_id in audit_to_row:
            # Update existing row
            row_idx = audit_to_row[audit_id]
            for col_offset, val in enumerate(row_values):
                ws.cell(row=row_idx, column=min_col + col_offset, value=val)
            updated += 1
        else:
            # Append new row after the last row of the table
            new_row = max_row + 1
            for col_offset, val in enumerate(row_values):
                ws.cell(row=new_row, column=min_col + col_offset, value=val)
            # Expand the table reference to include the new row
            new_ref = ref.split(":")[0] + ":" + \
                      f"{chr(ord('A') + min_col - 1 + len(row_values) - 1)}{new_row}"
            # Use openpyxl table ref update
            from openpyxl.utils import get_column_letter
            end_col = get_column_letter(min_col + len(row_values) - 1)
            tbl.ref = f"{ref.split(':')[0]}:{end_col}{new_row}"
            max_row = new_row
            audit_to_row[audit_id] = new_row
            inserted += 1

    return inserted, updated


# ── COMPLIANCE CALCULATION ───────────────────────────────────────────────────
def build_compliance(all_records):
    """
    Given the full debrief history, calculate rolling 7-day compliance per tail.
    Returns list of plane dicts for crosswinds_data.json.
    """
    today = date.today()
    window_start = today - timedelta(days=WINDOW_DAYS - 1)

    # Group by tail, sorted newest-first
    by_tail = {}
    for rec in all_records:
        tail = (str(rec.get("Tail") or "")).strip().upper()
        if tail not in TAILS:
            continue
        raw_date = rec.get("Date")
        if raw_date is None:
            continue
        if isinstance(raw_date, datetime):
            d = raw_date.date()
        elif isinstance(raw_date, date):
            d = raw_date
        else:
            try:
                d = datetime.fromisoformat(str(raw_date)).date()
            except Exception:
                continue

        if tail not in by_tail:
            by_tail[tail] = []
        by_tail[tail].append({
            "date":        d,
            "location":    str(rec.get("Location") or "").strip(),
            "tech":        str(rec.get("Name") or "").strip(),
            "audit_id":    str(rec.get("Audit ID") or "").strip(),
            "template_id": str(rec.get("Template ID") or "").strip(),
            "report_url":  str(rec.get("Report Link") or "").strip(),
        })

    for tail in by_tail:
        by_tail[tail].sort(key=lambda x: x["date"], reverse=True)

    planes = []
    for tail in TAILS:
        records = by_tail.get(tail, [])

        # Count cleans in last 7 days
        recent_7d = [r for r in records if r["date"] >= window_start]
        count_7d  = len(recent_7d)

        last_clean    = records[0]["date"].isoformat() if records else None
        last_location = records[0]["location"] if records else None
        last_tech     = records[0]["tech"]     if records else None

        days_since = (today - date.fromisoformat(last_clean)).days if last_clean else None

        # Status
        if count_7d >= 2:
            status = "compliant"
        elif count_7d == 1:
            oldest_recent_days = (today - recent_7d[-1]["date"]).days
            status = "soon" if oldest_recent_days >= (WINDOW_DAYS - SOON_DAYS) else "compliant"
        else:
            status = "noncompliant"

        # Recent inspections for detail panel
        recent_inspections = [
            {
                "audit_id":   r["audit_id"],
                "date":       r["date"].isoformat(),
                "location":   r["location"],
                "tech":       r["tech"],
                "report_url": r["report_url"],
            }
            for r in records[:SHOW_HISTORY]
        ]

        planes.append({
            "tail":               tail,
            "status":             status,
            "count7d":            count_7d,
            "daysSinceLast":      days_since,
            "lastClean":          last_clean,
            "lastLocation":       last_location,
            "lastTech":           last_tech,
            "recentInspections":  recent_inspections,
        })

    return planes


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=== Crosswinds Compliance Data Relay ===")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")

    # Parse payload from Power Automate
    try:
        incoming = json.loads(os.environ["PAYLOAD"])
    except Exception as e:
        print(f"ERROR: Failed to parse PAYLOAD: {e}")
        sys.exit(1)

    print(f"Incoming inspections from Power Automate: {len(incoming)}")

    # Filter to known tails only
    incoming = [i for i in incoming if (i.get("tail") or "").strip().upper() in TAILS]
    print(f"  After tail validation: {len(incoming)}")

    # ── Step 1: Download Excel ────────────────────────────────────────────────
    token   = get_token()
    buffer  = download_excel(token)
    wb      = load_workbook(buffer)
    ws, tbl = find_table_sheet(wb)

    existing, *_ = read_table(ws, tbl)
    print(f"\nExisting records in SharePoint: {len(existing)}")

    # ── Step 2: Upsert incoming records ──────────────────────────────────────
    print("\nUpserting records...")
    inserted, updated = upsert_records(ws, tbl, existing, incoming)
    print(f"  Inserted: {inserted}  |  Updated: {updated}")

    # ── Step 3: Upload updated Excel ─────────────────────────────────────────
    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    # Re-acquire token in case of long run
    token = get_token()
    upload_excel(token, out_buffer)

    # ── Step 4: Read full table for compliance calculation ───────────────────
    print("\nReading full history for compliance calculation...")
    # Re-read from the in-memory workbook (already updated)
    wb2      = load_workbook(io.BytesIO(out_buffer.getvalue()))
    ws2, tbl2 = find_table_sheet(wb2)
    all_records_dict, *_ = read_table(ws2, tbl2)
    all_records = list(all_records_dict.values())
    print(f"  Total records: {len(all_records)}")

    # ── Step 5: Build compliance and write JSON ───────────────────────────────
    print("\nCalculating compliance...")
    planes = build_compliance(all_records)

    nc   = sum(1 for p in planes if p["status"] == "noncompliant")
    soon = sum(1 for p in planes if p["status"] == "soon")
    ok   = sum(1 for p in planes if p["status"] == "compliant")
    print(f"  Compliant: {ok}  |  Due soon: {soon}  |  Noncompliant: {nc}")

    output = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "planes":    planes,
    }

    with open("crosswinds_data.json", "w") as f:
        json.dump(output, f, indent=2)

    print(f"\nWritten: crosswinds_data.json ({len(planes)} planes)")
    print("=== Done ===")
