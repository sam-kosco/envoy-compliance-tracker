"""
Envoy Compliance Data Relay
============================
Reads Envoy_Debriefs.xlsx from SharePoint, builds the combined debrief table
from both JotForm sheets, calculates compliance windows for each tail in the
Tail List, and writes data.json for the GitHub Pages dashboard.

Sheets read:
  Tail List     — master list of 187 Envoy tail numbers (column A)
  Envoy General — all debrief rows from the general JotForm
  DFW           — debrief rows filtered to Envoy Aircraft == Yes

Compliance jobs:
  ED1  — Exterior Detail #1  — 30-day cycle
  ED2  — Exterior Detail #2  — 60-day cycle

IHC (Interior Heavy Clean) is surfaced as informational (last date shown)
but does NOT drive compliance windows.

Credentials required (GitHub Secrets):
  TENANT_ID       — Azure AD tenant ID
  CLIENT_ID       — Foxtrot Report Automation app ID
  CLIENT_SECRET   — Foxtrot Report Automation client secret
"""

import os
import json
import sys
import requests
from datetime import datetime, timezone, date

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

DRIVE_ID  = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"
FILE_PATH = "Power Flows/Debriefs/Envoy Debriefs.xlsx"

# Compliance cycle lengths in days
CYCLES = {"ED1": 30, "ED2": 60}

SOON_DAYS = 7


# ─────────────────────────────────────────────
# STEP 1: Get Graph API token
# ─────────────────────────────────────────────

def get_token():
    print("Acquiring Graph API token...")
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    token = resp.json()["access_token"]
    print("  Token acquired.")
    return token


# ─────────────────────────────────────────────
# STEP 2: Download Excel from SharePoint
# ─────────────────────────────────────────────

def download_excel(token):
    print(f"Downloading: {FILE_PATH}")
    encoded = FILE_PATH.replace(" ", "%20")
    url = (
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}"
        f"/root:/{encoded}:/content"
    )
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    resp.raise_for_status()
    path = "/tmp/envoy_debriefs.xlsx"
    with open(path, "wb") as f:
        f.write(resp.content)
    print(f"  Downloaded {len(resp.content):,} bytes → {path}")
    return path


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def flag(value) -> int:
    """0 if No/blank/null, 1 for anything else."""
    if value is None:
        return 0
    s = str(value).strip().lower()
    return 0 if s in ("no", "0", "", "none", "nan") else 1


def parse_date(value):
    """Return a date object or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def fmt_date(d):
    """Format date as ISO string for JSON."""
    return d.isoformat() if d else None


# ─────────────────────────────────────────────
# STEP 3: Read Tail List
# ─────────────────────────────────────────────

def read_tail_list(wb):
    ws = wb["Tail List"]
    tails = []
    for row in ws.iter_rows(values_only=True):
        val = row[0]
        if val and str(val).strip().upper() != "TAILS":
            tails.append(str(val).strip().upper())
    print(f"  Tail List: {len(tails)} tails")
    return tails


# ─────────────────────────────────────────────
# STEP 4: Parse Envoy General sheet
# ─────────────────────────────────────────────

def parse_general(wb):
    """
    Columns (0-indexed):
      0: Date  1: Name  2: Location  3: Tail Number
      4: IHC   5: ED1   6: ED2       7: Sub ID
    """
    ws = wb["Envoy General"]
    rows = list(ws.iter_rows(values_only=True))
    debriefs = []
    for row in rows[1:]:
        if not row[0]:
            continue
        debriefs.append({
            "date":     parse_date(row[0]),
            "name":     str(row[1] or "").strip(),
            "location": str(row[2] or "").strip(),
            "tail":     str(row[3] or "").strip().upper(),
            "IHC":      flag(row[4]),
            "ED1":      flag(row[5]),
            "ED2":      flag(row[6]),
        })
    print(f"  Envoy General: {len(debriefs)} rows")
    return debriefs


# ─────────────────────────────────────────────
# STEP 5: Parse DFW sheet (Envoy Aircraft = Yes only)
# ─────────────────────────────────────────────

def parse_dfw(wb):
    """
    Columns (0-indexed):
      0: Date  1: Name  2: Tail  3: Envoy Aircraft  4: Regional Carrier
      5: IHC   6: RRON  7: ED1   8: ED2             9: Sub ID
    """
    ws = wb["DFW"]
    rows = list(ws.iter_rows(values_only=True))
    debriefs = []
    skipped = 0
    for row in rows[1:]:
        if not row[0]:
            continue
        if str(row[3] or "").strip().lower() != "yes":
            skipped += 1
            continue
        debriefs.append({
            "date":     parse_date(row[0]),
            "name":     str(row[1] or "").strip(),
            "location": "DFW",
            "tail":     str(row[2] or "").strip().upper(),
            "IHC":      flag(row[5]),
            "ED1":      flag(row[7]),
            "ED2":      flag(row[8]),
        })
    print(f"  DFW sheet:     {len(debriefs)} rows kept, {skipped} non-Envoy skipped")
    return debriefs


# ─────────────────────────────────────────────
# STEP 6: Build planes compliance table
# ─────────────────────────────────────────────

def build_planes(tails, all_debriefs):
    """
    For each tail in the Tail List, find the most recent date on which
    each service (ED1, ED2, IHC) was performed (flag == 1), then
    calculate compliance windows.
    """
    today = date.today()

    # Group debriefs by tail
    by_tail = {}
    for d in all_debriefs:
        t = d["tail"]
        if t not in by_tail:
            by_tail[t] = []
        by_tail[t].append(d)

    planes = []
    for tail in tails:
        records = by_tail.get(tail, [])

        # Most recent date each service was performed (flag == 1)
        last = {"ED1": None, "ED2": None, "IHC": None}
        last_service = None  # most recent debrief of any kind

        for d in records:
            d_date = d["date"]
            if d_date is None:
                continue
            if last_service is None or d_date > last_service:
                last_service = d_date
            for job in ("ED1", "ED2", "IHC"):
                if d[job] == 1:
                    if last[job] is None or d_date > last[job]:
                        last[job] = d_date

        # Calculate compliance windows for tracked jobs
        windows = {}
        for job, cycle in CYCLES.items():
            if last[job] is None:
                windows[job] = "No Service"
            else:
                windows[job] = cycle - (today - last[job]).days

        # Most recent location
        sorted_records = sorted(
            [r for r in records if r["date"]],
            key=lambda r: r["date"],
            reverse=True
        )
        last_location = sorted_records[0]["location"] if sorted_records else None
        last_name     = sorted_records[0]["name"]     if sorted_records else None

        planes.append({
            "tail":        tail,
            "lastService": fmt_date(last_service),
            "lastED1":     fmt_date(last["ED1"]),
            "lastED2":     fmt_date(last["ED2"]),
            "lastIHC":     fmt_date(last["IHC"]),
            "lastLocation": last_location,
            "lastTech":     last_name,
            "ED1":         windows["ED1"],
            "ED2":         windows["ED2"],
        })

    print(f"  Built compliance for {len(planes)} planes")
    return planes


# ─────────────────────────────────────────────
# STEP 7: Format debriefs for JSON output
# ─────────────────────────────────────────────

def format_debriefs(all_debriefs):
    """Convert date objects to strings and return sorted newest-first."""
    out = []
    for d in all_debriefs:
        out.append({
            "tail":     d["tail"],
            "date":     fmt_date(d["date"]),
            "name":     d["name"],
            "location": d["location"],
            "IHC":      d["IHC"],
            "ED1":      d["ED1"],
            "ED2":      d["ED2"],
        })
    # Sort newest first
    out.sort(key=lambda r: r["date"] or "", reverse=True)
    return out


# ─────────────────────────────────────────────
# STEP 8: Write data.json
# ─────────────────────────────────────────────

def write_json(planes, debriefs):
    output = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "planes":    planes,
        "debriefs":  debriefs,
    }
    with open("data.json", "w") as f:
        json.dump(output, f, indent=2, default=str)
    print(f"  Written: data.json  ({len(planes)} planes, {len(debriefs)} debriefs)")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl requests")
        sys.exit(1)

    print("=== Envoy Compliance Data Relay ===")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")

    token     = get_token()
    xlsx_path = download_excel(token)

    print("\nParsing workbook...")
    import openpyxl as ox
    wb = ox.load_workbook(xlsx_path, data_only=True)

    tails        = read_tail_list(wb)
    general_rows = parse_general(wb)
    dfw_rows     = parse_dfw(wb)
    all_debriefs = general_rows + dfw_rows
    print(f"  Total combined debriefs: {len(all_debriefs)}")

    print("\nBuilding compliance table...")
    planes   = build_planes(tails, all_debriefs)
    debriefs = format_debriefs(all_debriefs)

    print("\nWriting data.json...")
    write_json(planes, debriefs)

    # Quick summary
    nc   = sum(1 for p in planes if any(
        isinstance(p[j], str) or (isinstance(p[j], int) and p[j] < 0)
        for j in ("ED1", "ED2")
    ))
    soon = sum(1 for p in planes if not any(
        isinstance(p[j], str) or (isinstance(p[j], int) and p[j] < 0)
        for j in ("ED1", "ED2")
    ) and any(
        isinstance(p[j], int) and 0 <= p[j] <= SOON_DAYS
        for j in ("ED1", "ED2")
    ))
    print(f"\n  Noncompliant: {nc}  |  Due soon: {soon}  |  OK: {len(planes)-nc-soon}")
    print("\n=== Done ===")
