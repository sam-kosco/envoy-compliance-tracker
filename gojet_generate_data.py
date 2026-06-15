"""
GoJet Compliance Data Relay
===========================
Reads "GoJet Debriefs.xlsx" from SharePoint, calculates compliance windows
for every tail on the Tails sheet, and writes gojet_data.json for the
GitHub Pages dashboard.

Tracked jobs (fixed 60-day cycles):
  ED1 — Exterior Detail #1 — 60-day cycle
  ED2 — Exterior Detail #2 — 60-day cycle
  CE  — Carpet Extraction  — 60-day cycle

Informational only (shown in detail panel, no compliance window):
  IHC — Interior Heavy Clean
  RON — RON Clean

Notes / differences vs Mesa:
  - The GoJet "Input" sheet HAS a Location column, so the dashboard shows
    last-location (like Envoy & PSA).
  - Debrief cells encode "Yes. <model>"; the trailing text is not used for
    compliance (same flag() logic as Mesa).
  - Tail numbers are bare ("501", "534", ...).

Local generation / testing:
  Set GOJET_LOCAL_XLSX=<path> to read a local workbook instead of SharePoint
  (skips Graph auth). Used to seed the initial committed gojet_data.json.

Credentials (GitHub Secrets — same as Envoy/PSA/Mesa trackers):
  TENANT_ID, CLIENT_ID, CLIENT_SECRET
"""

import os, json, sys
from datetime import datetime, timezone, date

DRIVE_ID  = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"
FILE_PATH = "Power Flows/Debriefs/GoJet Debriefs.xlsx"

CYCLES  = {"ED1": 60, "ED2": 60, "CE": 60}
TRACKED = ["ED1", "ED2", "CE"]
INFO    = ["IHC", "RON"]


def get_token():
    import requests
    print("Acquiring Graph API token...")
    resp = requests.post(
        f"https://login.microsoftonline.com/{os.environ['TENANT_ID']}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials", "client_id": os.environ["CLIENT_ID"],
              "client_secret": os.environ["CLIENT_SECRET"], "scope": "https://graph.microsoft.com/.default"}
    )
    resp.raise_for_status()
    print("  Token acquired.")
    return resp.json()["access_token"]


def download_excel(token):
    import requests
    print(f"Downloading: {FILE_PATH}")
    encoded = FILE_PATH.replace(" ", "%20")
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{encoded}:/content",
        headers={"Authorization": f"Bearer {token}"}
    )
    resp.raise_for_status()
    path = "/tmp/gojet_debriefs.xlsx"
    with open(path, "wb") as f:
        f.write(resp.content)
    print(f"  Downloaded {len(resp.content):,} bytes")
    return path


def flag(v):
    # Done cells are recorded as "Yes. <model>"; everything else (No, blank)
    # is not done. Detect on the word "yes".
    if v is None:
        return 0
    return 1 if "yes" in str(v).strip().lower() else 0


def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def fmt(d): return d.isoformat() if d else None


def parse_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # Tails roster — column 0 of the "Tails" sheet.
    ws_tl = wb["Tails"]
    tails = [str(r[0]).strip().upper() for r in ws_tl.iter_rows(values_only=True)
             if r[0] is not None and str(r[0]).strip().upper() not in ("", "TAIL NUMBER", "TAILS")]
    print(f"  Tails: {len(tails)} tails")

    # Debriefs — the "Input" sheet.
    # Cols: 0=Date,1=Name,2=Location,3=Tail,4=IHC,5=RON,6=ED1,7=ED2,8=CE,9=SubID
    ws_d = wb["Input"]
    debriefs = []
    for row in list(ws_d.iter_rows(values_only=True))[1:]:
        if not row or row[0] is None:
            continue
        debriefs.append({
            "date": parse_date(row[0]),
            "name": str(row[1] or "").strip(),
            "location": str(row[2] or "").strip().upper(),
            "tail": str(row[3] or "").strip().upper(),
            "IHC": flag(row[4]), "RON": flag(row[5]),
            "ED1": flag(row[6]), "ED2": flag(row[7]), "CE": flag(row[8]),
        })
    print(f"  Debriefs: {len(debriefs)} rows")
    return tails, debriefs


def build_planes(tails, debriefs):
    today = date.today()
    by_tail = {}
    for d in debriefs:
        by_tail.setdefault(d["tail"], []).append(d)

    planes = []
    for tail in tails:
        recs = by_tail.get(tail, [])
        last = {j: None for j in TRACKED + INFO}
        ls = None
        for d in recs:
            dd = d["date"]
            if dd is None: continue
            if ls is None or dd > ls: ls = dd
            for j in TRACKED + INFO:
                if d[j] == 1 and (last[j] is None or dd > last[j]):
                    last[j] = dd

        windows = {j: ("No Service" if last[j] is None else CYCLES[j] - (today - last[j]).days)
                   for j in TRACKED}

        sr = sorted([r for r in recs if r["date"]], key=lambda r: r["date"], reverse=True)
        planes.append({
            "tail": tail,
            "lastService": fmt(ls),
            "lastED1": fmt(last["ED1"]),
            "lastED2": fmt(last["ED2"]),
            "lastCE":  fmt(last["CE"]),
            "lastIHC": fmt(last["IHC"]),
            "lastRON": fmt(last["RON"]),
            "lastLocation": sr[0]["location"] if sr and sr[0]["location"] else None,
            "lastTech": sr[0]["name"] if sr else None,
            **windows,
        })

    print(f"  Built compliance for {len(planes)} planes")
    return planes


def format_debriefs(debriefs):
    out = []
    for d in debriefs:
        out.append({
            "tail": d["tail"], "date": fmt(d["date"]), "name": d["name"],
            "location": d["location"],
            "ED1": d["ED1"], "ED2": d["ED2"], "CE": d["CE"],
            "IHC": d["IHC"], "RON": d["RON"],
        })
    out.sort(key=lambda r: r["date"] or "", reverse=True)
    return out


if __name__ == "__main__":
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERROR: pip install openpyxl requests")
        sys.exit(1)

    print("=== GoJet Compliance Data Relay ===")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")

    local = os.environ.get("GOJET_LOCAL_XLSX")
    if local:
        print(f"Using local workbook: {local}")
        xlsx_path = local
    else:
        token     = get_token()
        xlsx_path = download_excel(token)

    print("\nParsing workbook...")
    tails, debriefs = parse_workbook(xlsx_path)

    print("\nBuilding compliance table...")
    planes       = build_planes(tails, debriefs)
    debriefs_out = format_debriefs(debriefs)

    output = {"generated": datetime.now(timezone.utc).isoformat(),
              "cycles": CYCLES, "planes": planes, "debriefs": debriefs_out}

    with open("gojet_data.json", "w") as f:
        json.dump(output, f, indent=2, default=str)

    nc   = sum(1 for p in planes if any(p[j] == "No Service" or
               (isinstance(p[j], int) and p[j] < 0) for j in TRACKED))
    soon = sum(1 for p in planes if not any(p[j] == "No Service" or
               (isinstance(p[j], int) and p[j] < 0) for j in TRACKED)
               and any(isinstance(p[j], int) and 0 <= p[j] <= 7 for j in TRACKED))

    print(f"\n  Written: gojet_data.json")
    print(f"  Noncompliant: {nc}  |  Due soon: {soon}  |  OK: {len(planes)-nc-soon}")
    print("\n=== Done ===")
