"""
Mesa Compliance Data Relay
==========================
Reads "Mesa Debriefs.xlsx" from SharePoint, calculates compliance windows
for every tail on the Tails sheet, and writes mesa_data.json for the
GitHub Pages dashboard.

Tracked jobs (cycles read from the "Service WIndow" sheet):
  IHC        — Interior Heavy Clean         — 45-day cycle
  ED         — Exterior Detail              — 30-day cycle
  DSC        — Deep Seat Clean              — 30-day cycle
  FD         — Detailed Flight Deck Clean   — 30-day cycle  ("Flight Deck")
  CE         — Carpet Extraction            — 30-day cycle

Informational only (shown in detail panel, no compliance window):
  RON — RON Clean
  EC  — Exterior Clean
  ESS — Disinfection

Ignored entirely (per program owner): FCD — Fleet Campaign Decal.

Notes / differences vs PSA & Envoy:
  - The Mesa Debriefs sheet has NO Location column, so there is no
    last-location data. Service cells encode "Yes. <number>"; the trailing
    number is not a location and is not used for compliance.
  - "Detailed Flight Deck Clean" lives in the LAST debriefs column, after Sub ID.

Credentials (GitHub Secrets — same as Envoy/PSA trackers):
  TENANT_ID, CLIENT_ID, CLIENT_SECRET
"""

import os, json, sys, requests
from datetime import datetime, timezone, date

TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

DRIVE_ID  = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"
FILE_PATH = "Power Flows/Debriefs/Mesa Debriefs.xlsx"

# Fallback cycles — overridden by whatever the "Service WIndow" sheet says.
DEFAULT_CYCLES = {"IHC": 45, "ED": 30, "DSC": 30, "FD": 30, "CE": 30}
TRACKED = ["IHC", "ED", "DSC", "FD", "CE"]
INFO    = ["RON", "EC", "ESS"]

# Maps the "Service" labels in the Service WIndow sheet to our job codes.
WINDOW_LABEL_TO_CODE = {
    "IHC": "IHC", "ED": "ED", "DSC": "DSC", "CE": "CE",
    "FLIGHT DECK": "FD", "FD": "FD", "DETAILED FLIGHT DECK CLEAN": "FD",
}


def get_token():
    print("Acquiring Graph API token...")
    resp = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials", "client_id": CLIENT_ID,
              "client_secret": CLIENT_SECRET, "scope": "https://graph.microsoft.com/.default"}
    )
    resp.raise_for_status()
    print("  Token acquired.")
    return resp.json()["access_token"]


def download_excel(token):
    print(f"Downloading: {FILE_PATH}")
    encoded = FILE_PATH.replace(" ", "%20")
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{encoded}:/content",
        headers={"Authorization": f"Bearer {token}"}
    )
    resp.raise_for_status()
    path = "/tmp/mesa_debriefs.xlsx"
    with open(path, "wb") as f:
        f.write(resp.content)
    print(f"  Downloaded {len(resp.content):,} bytes")
    return path


def flag(v):
    # Done cells are recorded as "Yes. <number>"; everything else (No, blank,
    # "No.", stray ":") is not done. Detect on the word "yes".
    if v is None:
        return 0
    return 1 if "yes" in str(v).strip().lower() else 0


def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def fmt(d): return d.isoformat() if d else None


def read_cycles(wb):
    """Read compliance windows from the 'Service WIndow' sheet, falling back to
    DEFAULT_CYCLES for any tracked job that is missing or unparseable."""
    cycles = dict(DEFAULT_CYCLES)
    sheet = next((wb[n] for n in wb.sheetnames if n.strip().lower() == "service window"), None)
    if sheet is None:
        print("  WARNING: 'Service WIndow' sheet not found — using default cycles.")
        return cycles
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().upper()
        code = WINDOW_LABEL_TO_CODE.get(label)
        if not code:
            continue
        try:
            cycles[code] = int(row[1])
        except (TypeError, ValueError):
            pass
    print(f"  Cycles: {cycles}")
    return cycles


def parse_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    cycles = read_cycles(wb)

    # Tails roster — column 0 of the "Tails" sheet.
    ws_tl = wb["Tails"]
    tails = [str(r[0]).strip().upper() for r in ws_tl.iter_rows(values_only=True)
             if r[0] and str(r[0]).strip().upper() != "TAILS"]
    print(f"  Tails: {len(tails)} tails")

    # Debriefs
    # Cols: 0=Date,1=Name,2=Tail,3=IHC,4=RON,5=EC,6=ED,7=DSC,8=CE,
    #       9=FCD(ignored),10=ESS,11=SubID,12=Flight Deck
    ws_d = wb["Debriefs"]
    debriefs = []
    for row in list(ws_d.iter_rows(values_only=True))[1:]:
        if not row or row[0] is None:
            continue
        debriefs.append({
            "date": parse_date(row[0]),
            "name": str(row[1] or "").strip(),
            "tail": str(row[2] or "").strip().upper(),
            "IHC": flag(row[3]),  "RON": flag(row[4]), "EC":  flag(row[5]),
            "ED":  flag(row[6]),  "DSC": flag(row[7]), "CE":  flag(row[8]),
            "ESS": flag(row[10]), "FD":  flag(row[12] if len(row) > 12 else None),
        })
    print(f"  Debriefs: {len(debriefs)} rows")
    return tails, debriefs, cycles


def build_planes(tails, debriefs, cycles):
    today = date.today()
    by_tail = {}
    for d in debriefs:
        by_tail.setdefault(d["tail"], []).append(d)

    planes = []
    for tail in tails:
        recs = by_tail.get(tail, [])
        last = {j: None for j in TRACKED + INFO}
        ls = None
        for d in recs:
            dd = d["date"]
            if dd is None: continue
            if ls is None or dd > ls: ls = dd
            for j in TRACKED + INFO:
                if d[j] == 1 and (last[j] is None or dd > last[j]):
                    last[j] = dd

        windows = {j: ("No Service" if last[j] is None else cycles[j] - (today - last[j]).days)
                   for j in TRACKED}

        sr = sorted([r for r in recs if r["date"]], key=lambda r: r["date"], reverse=True)
        planes.append({
            "tail": tail,
            "lastService": fmt(ls),
            "lastIHC": fmt(last["IHC"]),
            "lastED":  fmt(last["ED"]),
            "lastDSC": fmt(last["DSC"]),
            "lastFD":  fmt(last["FD"]),
            "lastCE":  fmt(last["CE"]),
            "lastRON": fmt(last["RON"]),
            "lastEC":  fmt(last["EC"]),
            "lastESS": fmt(last["ESS"]),
            "lastTech": sr[0]["name"] if sr else None,
            **windows,
        })

    print(f"  Built compliance for {len(planes)} planes")
    return planes


def format_debriefs(debriefs):
    out = []
    for d in debriefs:
        out.append({
            "tail": d["tail"], "date": fmt(d["date"]), "name": d["name"],
            "IHC": d["IHC"], "ED": d["ED"], "DSC": d["DSC"],
            "FD": d["FD"], "CE": d["CE"],
            "RON": d["RON"], "EC": d["EC"], "ESS": d["ESS"],
        })
    out.sort(key=lambda r: r["date"] or "", reverse=True)
    return out


if __name__ == "__main__":
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print("ERROR: pip install openpyxl requests")
        sys.exit(1)

    print("=== Mesa Compliance Data Relay ===")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")

    token     = get_token()
    xlsx_path = download_excel(token)

    print("\nParsing workbook...")
    tails, debriefs, cycles = parse_workbook(xlsx_path)

    print("\nBuilding compliance table...")
    planes       = build_planes(tails, debriefs, cycles)
    debriefs_out = format_debriefs(debriefs)

    output = {"generated": datetime.now(timezone.utc).isoformat(),
              "cycles": cycles, "planes": planes, "debriefs": debriefs_out}

    with open("mesa_data.json", "w") as f:
        json.dump(output, f, indent=2, default=str)

    nc   = sum(1 for p in planes if any(p[j] == "No Service" or
               (isinstance(p[j], int) and p[j] < 0) for j in TRACKED))
    soon = sum(1 for p in planes if not any(p[j] == "No Service" or
               (isinstance(p[j], int) and p[j] < 0) for j in TRACKED)
               and any(isinstance(p[j], int) and 0 <= p[j] <= 7 for j in TRACKED))

    print(f"\n  Written: mesa_data.json")
    print(f"  Noncompliant: {nc}  |  Due soon: {soon}  |  OK: {len(planes)-nc-soon}")
    print("\n=== Done ===")
