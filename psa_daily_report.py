"""
PSA Daily Compliance Report
===========================
Builds "PSA-FoxTrot Compliance MM-DD-YYYY.xlsx" — a snapshot styled after the
Tail List sheet of PSA Debriefs.xlsx — and emails it to EMAIL_LIST from
foxtrot.automation@foxtrotaviation.com via Microsoft Graph.

Reads psa_data.json (the workflow runs psa_generate_data.py immediately
before this script, so the data is fresh — no compliance logic is
duplicated here).

Each service date cell is colored by that service's compliance window:
  Green  — window > 7 days (compliant)
  Yellow — window 0-7 days (due soon)
  Red    — window < 0 or never serviced (noncompliant)
The Last Service cell is colored by the tail's overall status.

Runs on GitHub Actions (psa_daily_report.yml) daily at 9 AM Eastern.
The workflow crons at 13:00 and 14:00 UTC; the guard below sends only on
the run where it is 9 AM in America/New_York, year-round across DST.

Env:
  TENANT_ID / CLIENT_ID / CLIENT_SECRET — Graph credentials (GitHub Secrets)
  RECIPIENTS_OVERRIDE — comma-separated list replacing EMAIL_LIST (testing)
  DRY_RUN=1           — build the workbook but do not email it
"""

import os
import sys
import json
import base64
from datetime import datetime, date
from zoneinfo import ZoneInfo

# ─────────────────────────────────────────────
# RECIPIENTS — edit this list to change who gets the daily report
# ─────────────────────────────────────────────

EMAIL_LIST = [
    "samuel.kosco@foxtrotaviation.com",
    "daniel.digiambattista@foxtrotaviation.com",
    "brad.decker@PSAAirlines.com",
    "daniel.starcher@psaairlines.com",
    "Matthew.Key@PSAAirlines.com",
]

SENDER = "foxtrot.automation@foxtrotaviation.com"

# Tracked services, in Tail List column order
SERVICES  = ["CC", "DSC", "CE", "ED1", "ED2", "Lav"]
SOON_DAYS = 7

# Classic Excel conditional-format palettes (fill, font)
GREEN  = ("C6EFCE", "006100")
YELLOW = ("FFEB9C", "9C6500")
RED    = ("FFC7CE", "9C0006")


def eastern_today():
    return datetime.now(ZoneInfo("America/New_York")).date()


def nine_am_guard():
    """On scheduled runs, proceed only when it is 9 AM Eastern (the workflow
    fires at both 13:00 and 14:00 UTC to cover EDT and EST)."""
    if os.environ.get("GITHUB_EVENT_NAME") != "schedule":
        return
    hour = datetime.now(ZoneInfo("America/New_York")).hour
    if hour != 9:
        print(f"Not 9 AM Eastern (hour={hour}) — skipping this scheduled run.")
        sys.exit(0)


def classify(window):
    """Map a compliance window to a status color pair."""
    if isinstance(window, str) or window < 0:   # "No Service" or overdue
        return RED
    if window <= SOON_DAYS:
        return YELLOW
    return GREEN


def overall(plane):
    """Tail-level status across all tracked services (tracker rules)."""
    windows = [plane[s] for s in SERVICES]
    if any(isinstance(w, str) or w < 0 for w in windows):
        return RED
    if any(w <= SOON_DAYS for w in windows):
        return YELLOW
    return GREEN


# ─────────────────────────────────────────────
# Build the workbook
# ─────────────────────────────────────────────

def build_workbook(planes, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tail List"

    headers = ["Tails", "Last Service", "Last CC", "Last DSC", "Last CE",
               "Last ED1", "Last ED2", "Last Lav"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin   = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center, border

    def paint(cell, iso_date, color):
        fill, font = color
        if iso_date:
            cell.value = datetime.strptime(iso_date, "%Y-%m-%d").date()
            cell.number_format = "mm/dd/yyyy"
        else:
            cell.value = "No Service"
        cell.fill      = PatternFill("solid", fgColor=fill)
        cell.font      = Font(color=font)
        cell.alignment = center
        cell.border    = border

    for r, p in enumerate(planes, start=2):
        tc = ws.cell(row=r, column=1, value=p["tail"])
        tc.font, tc.alignment, tc.border = Font(bold=True), center, border
        paint(ws.cell(row=r, column=2), p.get("lastService"), overall(p))
        for i, svc in enumerate(SERVICES, start=3):
            paint(ws.cell(row=r, column=i), p.get(f"last{svc}"), classify(p[svc]))

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14
    ws.auto_filter.ref = f"A1:H{len(planes) + 1}"

    wb.save(out_path)
    print(f"  Workbook written: {out_path} ({len(planes)} tails)")


# ─────────────────────────────────────────────
# Email via Microsoft Graph
# ─────────────────────────────────────────────

def get_token():
    import requests
    resp = requests.post(
        f"https://login.microsoftonline.com/{os.environ['TENANT_ID']}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials",
              "client_id": os.environ["CLIENT_ID"],
              "client_secret": os.environ["CLIENT_SECRET"],
              "scope": "https://graph.microsoft.com/.default"},
    )
    resp.raise_for_status()
    return resp.json()["access_token"]


def send_report(token, recipients, xlsx_path, filename, report_date, counts):
    import requests
    with open(xlsx_path, "rb") as f:
        payload = base64.b64encode(f.read()).decode()

    nc, soon, ok = counts
    body = f"""
<p>Attached is the PSA-FoxTrot compliance report for <strong>{report_date.strftime('%B %d, %Y')}</strong>.</p>
<table style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;margin-top:12px">
  <tr><td style="padding:6px 14px;font-weight:bold;color:#9C0006;background:#FFC7CE">Noncompliant</td><td style="padding:6px 14px">{nc}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#9C6500;background:#FFEB9C">Due Soon</td><td style="padding:6px 14px">{soon}</td></tr>
  <tr><td style="padding:6px 14px;font-weight:bold;color:#006100;background:#C6EFCE">Compliant</td><td style="padding:6px 14px">{ok}</td></tr>
</table>
<p style="margin-top:16px">Date cells are colored per service: green = compliant, yellow = due within {SOON_DAYS} days, red = overdue or never serviced.</p>
<p style="margin-top:16px;color:#888;font-size:12px">— Foxtrot Aviation Services PSA Compliance Tracker</p>
"""
    resp = requests.post(
        f"https://graph.microsoft.com/v1.0/users/{SENDER}/sendMail",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"message": {
            "subject": f"PSA-FoxTrot Compliance {report_date.strftime('%m/%d/%Y')}",
            "body": {"contentType": "HTML", "content": body},
            "toRecipients": [{"emailAddress": {"address": a}} for a in recipients],
            "attachments": [{
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": filename,
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "contentBytes": payload,
            }],
        }},
    )
    resp.raise_for_status()
    print(f"  Email sent to: {', '.join(recipients)}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=== PSA Daily Compliance Report ===")
    nine_am_guard()

    report_date = eastern_today()
    # Slashes are invalid in filenames, so the attachment uses dashes
    filename = f"PSA-FoxTrot Compliance {report_date.strftime('%m-%d-%Y')}.xlsx"
    out_path = f"/tmp/{filename}" if os.name != "nt" else filename

    with open("psa_data.json") as f:
        planes = json.load(f)["planes"]
    print(f"  Loaded psa_data.json: {len(planes)} tails")

    counts = (
        sum(1 for p in planes if overall(p) == RED),
        sum(1 for p in planes if overall(p) == YELLOW),
        sum(1 for p in planes if overall(p) == GREEN),
    )
    print(f"  Noncompliant: {counts[0]}  |  Due soon: {counts[1]}  |  OK: {counts[2]}")

    build_workbook(planes, out_path)

    if os.environ.get("DRY_RUN") == "1":
        print("  DRY_RUN=1 — skipping email.")
        sys.exit(0)

    override = os.environ.get("RECIPIENTS_OVERRIDE", "").strip()
    recipients = [a.strip() for a in override.split(",") if a.strip()] or EMAIL_LIST

    token = get_token()
    send_report(token, recipients, out_path, filename, report_date, counts)
    print("=== Done ===")
