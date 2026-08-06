"""
PSA Compliance Data Relay
==========================
Reads PSA_Debriefs.xlsx from SharePoint, calculates compliance windows
for all 155 tails, and writes psa_data.json for the GitHub Pages dashboard.

Tracked jobs (all from Debriefs sheet):
  CC  — Cockpit Clean            — 30-day cycle
  DSC — Deep Seat Clean          — 30-day cycle
  CE  — Carpet Extraction        — 30-day cycle
  ED1 — Exterior Detail #1       — 30-day cycle
  ED2 — Exterior Detail #2       — 30-day cycle
  Lav — Lav Tank Pressure Wash   — 90-day cycle

Informational only (shown in detail panel, no compliance window):
  IC  — Interior Clean
  EC  — Exterior Clean
  ED3 — Exterior Detail #3
  ED4 — Exterior Detail #4

Credentials (GitHub Secrets — same as Envoy tracker):
  TENANT_ID, CLIENT_ID, CLIENT_SECRET
"""

import os, json, sys, requests
from datetime import datetime, timezone, date

TENANT_ID     = os.environ["TENANT_ID"]
CLIENT_ID     = os.environ["CLIENT_ID"]
CLIENT_SECRET = os.environ["CLIENT_SECRET"]

DRIVE_ID  = "b!_bzXaIx86kOufgJN3ih-BaDIDthKYuxJkJtLi1Bm5irGjCEnK-VHSpBRRm3_SDKU"
FILE_PATH = "Power Flows/Debriefs/PSA Debriefs.xlsx"

CYCLES  = {"CC": 30, "DSC": 30, "CE": 30, "ED1": 30, "ED2": 30, "Lav": 90}
TRACKED = list(CYCLES.keys())
INFO    = ["IC", "EC", "ED3", "ED4"]

# ── SafetyCulture audit integration ──────────────────────────────────────────
# The "Audits" sheet in PSA Debriefs.xlsx logs SafetyCulture inspections pulled
# per PSA service. Each inspection is matched to the debrief for the job it
# audited and its public web report is linked in the dashboard service history.
#
# Match key: Tail + Location + Service, plus Date within a small tolerance.
# Empirically the SafetyCulture "Service Date" runs exactly ONE day after the
# debrief date (the crew debriefs on the service night; the audit is logged the
# next day). We match the nearest service-performed debrief within the window
# below rather than requiring an exact date so the link survives future drift.
# Multiple inspections can attach to one debrief (audits are per individual job).
#
# Only DSC is wired up in SafetyCulture today; add rows as more services go live.
AUDIT_SERVICE_MAP    = {"DSC": "DSC", "CC": "CC", "CE": "CE",
                        "ED1": "ED1", "ED2": "ED2", "LAV": "Lav"}
AUDIT_DATE_TOLERANCE = 3  # days, +/- around the debrief date


def get_token():
    print("Acquiring Graph API token...")
    resp = requests.post(
        f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token",
        data={"grant_type": "client_credentials", "client_id": CLIENT_ID,
              "client_secret": CLIENT_SECRET, "scope": "https://graph.microsoft.com/.default"}
    )
    resp.raise_for_status()
    print("  Token acquired.")
    return resp.json()["access_token"]


def download_excel(token):
    print(f"Downloading: {FILE_PATH}")
    encoded = FILE_PATH.replace(" ", "%20")
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/root:/{encoded}:/content",
        headers={"Authorization": f"Bearer {token}"}
    )
    resp.raise_for_status()
    path = "/tmp/psa_debriefs.xlsx"
    with open(path, "wb") as f:
        f.write(resp.content)
    print(f"  Downloaded {len(resp.content):,} bytes")
    return path


def flag(v):
    if v is None: return 0
    return 0 if str(v).strip().lower() in ("no", "0", "", "none", "nan") else 1


def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def fmt(d): return d.isoformat() if d else None


def parse_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # Tails — column A; Status in column I ("Disabled" hides the tail from
    # the tracker; anything else, including blank, shows it — so a missing
    # status on a newly added row doesn't silently drop the tail).
    # Duplicate rows are deduped.
    ws_tl = wb["Tail List"]
    tails, seen, disabled = [], set(), 0
    for r in ws_tl.iter_rows(values_only=True):
        t = str(r[0] or "").strip().upper()
        if not t or t == "TAILS" or t in seen:
            continue
        seen.add(t)
        status = str(r[8] or "").strip().lower() if len(r) > 8 else ""
        if status == "disabled":
            disabled += 1
            continue
        tails.append(t)
    print(f"  Tail List: {len(tails)} active tails ({disabled} disabled)")

    # Debriefs
    # Cols: 0=Date,1=Name,2=Location,3=Tail,4=IC,5=EC,6=CC,7=DSC,8=CE,
    #       9=ED1,10=ED2,11=ED3,12=ED4,13=Lav,14=SubID
    ws_d = wb["Debriefs"]
    debriefs = []
    for row in list(ws_d.iter_rows(values_only=True))[1:]:
        if not row[0]: continue
        loc = str(row[2] or "").strip()
        loc = loc.replace("-PSA", "").replace("-psa", "").strip()
        debriefs.append({
            "date":     parse_date(row[0]),
            "name":     str(row[1] or "").strip(),
            "location": loc,
            "tail":     str(row[3] or "").strip().upper(),
            "IC":  flag(row[4]),  "EC":  flag(row[5]),
            "CC":  flag(row[6]),  "DSC": flag(row[7]),
            "CE":  flag(row[8]),  "ED1": flag(row[9]),
            "ED2": flag(row[10]), "ED3": flag(row[11]),
            "ED4": flag(row[12]), "Lav": flag(row[13]),
        })
    print(f"  Debriefs: {len(debriefs)} rows")

    # Audits (SafetyCulture inspections)
    # Cols: 0=Submitter,1=Audit id,2=Location,3=Tail,4=Service Date,
    #       5=Service,6=Web Report Link,7=View Report,8=Status,9=Approval Date
    audits = []
    if "Audits" in wb.sheetnames:
        ws_a = wb["Audits"]
        for row in list(ws_a.iter_rows(values_only=True))[1:]:
            if not row or not row[1]: continue          # need an Audit id
            url = str(row[6] or "").strip()
            if not url: continue                          # no report to link
            loc = str(row[2] or "").strip()
            loc = loc.replace("-PSA", "").replace("-psa", "").strip()
            audits.append({
                "auditId":   str(row[1]).strip(),
                "submitter": str(row[0] or "").strip(),
                "location":  loc,
                "tail":      str(row[3] or "").strip().upper(),
                "date":      parse_date(row[4]),
                "service":   str(row[5] or "").strip(),
                "url":       url,
                "status":    str(row[8] or "").strip(),
            })
        print(f"  Audits: {len(audits)} inspections")
    else:
        print("  Audits: sheet not present — skipping")

    return tails, debriefs, audits


def attach_audits(debriefs, audits):
    """Match each SafetyCulture inspection to the debrief for the job it audited
    and attach the web report link to that debrief.

    A match requires the same Tail, same Location, and the audited Service to
    have been performed on the debrief (its flag == 1), with the debrief date
    within AUDIT_DATE_TOLERANCE days of the inspection's Service Date. The
    nearest such debrief wins (tie broken toward the debrief on/before the audit,
    which mirrors the observed one-day offset). Unmatched inspections are
    discarded; a single debrief may collect multiple inspections.
    """
    from collections import defaultdict
    idx = defaultdict(list)
    for d in debriefs:
        d["audits"] = []
        idx[(d["tail"], d["location"].upper())].append(d)

    matched = 0
    for a in audits:
        svc_key = AUDIT_SERVICE_MAP.get(a["service"].strip().upper())
        if not svc_key or a["date"] is None:
            continue
        cands = [d for d in idx.get((a["tail"], a["location"].upper()), [])
                 if d["date"] is not None and d.get(svc_key) == 1
                 and abs((d["date"] - a["date"]).days) <= AUDIT_DATE_TOLERANCE]
        if not cands:
            continue
        best = min(cands, key=lambda d: (abs((d["date"] - a["date"]).days),
                                         0 if d["date"] <= a["date"] else 1))
        best["audits"].append({
            "service":   svc_key,
            "url":       a["url"],
            "auditId":   a["auditId"],
            "status":    a["status"],
            "submitter": a["submitter"],
            "date":      fmt(a["date"]),
        })
        matched += 1

    print(f"  Audits matched to debriefs: {matched}/{len(audits)} "
          f"({len(audits) - matched} discarded)")
    return matched


def build_planes(tails, debriefs):
    today = date.today()
    by_tail = {}
    for d in debriefs:
        t = d["tail"]
        if t not in by_tail: by_tail[t] = []
        by_tail[t].append(d)

    planes = []
    for tail in tails:
        recs = by_tail.get(tail, [])
        last = {j: None for j in TRACKED + INFO}
        ls = None
        for d in recs:
            dd = d["date"]
            if dd is None: continue
            if ls is None or dd > ls: ls = dd
            for j in TRACKED + INFO:
                if d[j] == 1 and (last[j] is None or dd > last[j]):
                    last[j] = dd

        windows = {j: ("No Service" if last[j] is None else CYCLES[j] - (today - last[j]).days)
                   for j in TRACKED}

        sr = sorted([r for r in recs if r["date"]], key=lambda r: r["date"], reverse=True)
        planes.append({
            "tail": tail,
            "lastService":  fmt(ls),
            "lastCC":       fmt(last["CC"]),
            "lastDSC":      fmt(last["DSC"]),
            "lastCE":       fmt(last["CE"]),
            "lastED1":      fmt(last["ED1"]),
            "lastED2":      fmt(last["ED2"]),
            "lastLav":      fmt(last["Lav"]),
            "lastIC":       fmt(last["IC"]),
            "lastEC":       fmt(last["EC"]),
            "lastED3":      fmt(last["ED3"]),
            "lastED4":      fmt(last["ED4"]),
            "lastLocation": sr[0]["location"] if sr else None,
            "lastTech":     sr[0]["name"]     if sr else None,
            **windows,
        })

    print(f"  Built compliance for {len(planes)} planes")
    return planes


def format_debriefs(debriefs):
    out = []
    for d in debriefs:
        out.append({
            "tail": d["tail"], "date": fmt(d["date"]),
            "name": d["name"], "location": d["location"],
            "IC": d["IC"],  "EC": d["EC"],  "CC": d["CC"],
            "DSC": d["DSC"], "CE": d["CE"], "ED1": d["ED1"],
            "ED2": d["ED2"], "ED3": d["ED3"], "ED4": d["ED4"],
            "Lav": d["Lav"],
            "audits": d.get("audits", []),
        })
    out.sort(key=lambda r: r["date"] or "", reverse=True)
    return out


if __name__ == "__main__":
    try:
        import openpyxl
    except ImportError:
        print("ERROR: pip install openpyxl requests")
        sys.exit(1)

    print("=== PSA Compliance Data Relay ===")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n")

    token     = get_token()
    xlsx_path = download_excel(token)

    print("\nParsing workbook...")
    tails, debriefs, audits = parse_workbook(xlsx_path)

    print("\nMatching SafetyCulture audits to debriefs...")
    attach_audits(debriefs, audits)

    print("\nBuilding compliance table...")
    planes          = build_planes(tails, debriefs)
    debriefs_out    = format_debriefs(debriefs)

    output = {"generated": datetime.now(timezone.utc).isoformat(),
              "planes": planes, "debriefs": debriefs_out}

    with open("psa_data.json", "w") as f:
        json.dump(output, f, indent=2, default=str)

    nc   = sum(1 for p in planes if any(p[j] == "No Service" or
               (isinstance(p[j], int) and p[j] < 0) for j in TRACKED))
    soon = sum(1 for p in planes if not any(p[j] == "No Service" or
               (isinstance(p[j], int) and p[j] < 0) for j in TRACKED)
               and any(isinstance(p[j], int) and 0 <= p[j] <= 7 for j in TRACKED))

    print(f"\n  Written: psa_data.json")
    print(f"  Noncompliant: {nc}  |  Due soon: {soon}  |  OK: {len(planes)-nc-soon}")
    print("\n=== Done ===")
